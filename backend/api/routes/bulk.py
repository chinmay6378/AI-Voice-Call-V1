"""
Bulk campaign endpoints.

POST /bulk/upload              — parse CSV/Excel, create campaign (pending)
POST /bulk/campaigns/{id}/start — start calling
POST /bulk/campaigns/{id}/stop  — stop campaign
GET  /bulk/campaigns            — list all campaigns
GET  /bulk/campaigns/{id}       — campaign detail + contact statuses
"""
from __future__ import annotations

import csv
import io
import uuid
from typing import Annotated

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from config.settings import Settings, get_settings
from database.repository import (
    create_campaign,
    get_call,
    get_campaign,
    get_campaign_contacts,
    get_session,
    list_campaigns,
    update_campaign_status,
)
from database.models.campaign import CampaignStatus
from database.schemas.campaign import (
    CampaignResponse,
    ContactIn,
    ContactStatusResponse,
    UploadPreviewResponse,
)
from services.campaign_runner import is_running, start_campaign, stop_campaign
from utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/bulk", tags=["bulk"])

_NAME_COLS = {"name", "customer name", "customer_name", "full name", "full_name", "contact"}
_PHONE_COLS = {"phone", "phone_number", "phone number", "mobile", "mobile number", "mobile_number", "number", "contact number"}


# ── File parsing ──────────────────────────────────────────────────────────────

def _parse_csv(content: bytes) -> tuple[list[ContactIn], list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], ["No columns found in CSV"]
    return _extract_contacts(reader.fieldnames, list(reader))


def _parse_excel(content: bytes) -> tuple[list[ContactIn], list[str]]:
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl not installed — run: pip install openpyxl"]

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["Empty spreadsheet"]

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = [dict(zip(headers, [str(v).strip() if v is not None else "" for v in row])) for row in rows[1:]]
    return _extract_contacts(headers, data_rows)


def _extract_contacts(
    headers: list[str], rows: list[dict]
) -> tuple[list[ContactIn], list[str]]:
    errors: list[str] = []
    lower_headers = {h.lower().strip(): h for h in headers}

    name_col = next((lower_headers[h] for h in lower_headers if h in _NAME_COLS), None)
    phone_col = next((lower_headers[h] for h in lower_headers if h in _PHONE_COLS), None)

    if not name_col:
        errors.append(f"Could not find a 'name' column. Found: {list(lower_headers.keys())}")
    if not phone_col:
        errors.append(f"Could not find a 'phone' column. Found: {list(lower_headers.keys())}")
    if errors:
        return [], errors

    contacts: list[ContactIn] = []
    for i, row in enumerate(rows, start=2):
        name = str(row.get(name_col, "")).strip()
        phone = str(row.get(phone_col, "")).strip()
        if not name and not phone:
            continue  # skip blank rows
        if not name:
            errors.append(f"Row {i}: missing name")
            continue
        if not phone:
            errors.append(f"Row {i}: missing phone for {name}")
            continue
        if not phone.startswith("+"):
            phone = "+" + phone
        contacts.append(ContactIn(name=name, phone_number=phone))

    return contacts, errors


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=UploadPreviewResponse, status_code=status.HTTP_201_CREATED)
async def upload_contacts(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session),
    settings: Settings = Depends(get_settings),
) -> UploadPreviewResponse:
    """
    Parse a CSV or Excel file and create a pending campaign.
    Returns the campaign ID and a preview of parsed contacts.
    """
    name = file.filename or "campaign"
    content = await file.read()

    fn = name.lower()
    if fn.endswith(".csv"):
        contacts, errors = _parse_csv(content)
    elif fn.endswith((".xlsx", ".xls")):
        contacts, errors = _parse_excel(content)
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx/.xls files are supported")

    if not contacts:
        raise HTTPException(
            status_code=422,
            detail={"message": "No valid contacts found", "errors": errors},
        )

    campaign = await create_campaign(
        session,
        name=name,
        contacts=[c.model_dump() for c in contacts],
    )

    logger.info("bulk.uploaded", campaign_id=campaign.id, total=len(contacts), errors=len(errors))

    return UploadPreviewResponse(
        campaign_id=campaign.id,
        name=name,
        total=len(contacts),
        contacts=contacts,
        errors=errors,
    )


@router.post("/campaigns/{campaign_id}/start", response_model=CampaignResponse)
async def start(
    campaign_id: str,
    session: AsyncSession = Depends(get_session),
    settings: Annotated[Settings, Depends(get_settings)] = None,
) -> CampaignResponse:
    campaign = await get_campaign(session, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status == CampaignStatus.RUNNING:
        raise HTTPException(status_code=409, detail="Campaign is already running")
    if campaign.status == CampaignStatus.COMPLETED:
        raise HTTPException(status_code=409, detail="Campaign already completed")

    await start_campaign(campaign_id, settings)
    await session.refresh(campaign)

    contacts = await get_campaign_contacts(session, campaign_id)
    return _to_response(campaign, contacts)


@router.post("/campaigns/{campaign_id}/stop", response_model=CampaignResponse)
async def stop(
    campaign_id: str,
    session: AsyncSession = Depends(get_session),
) -> CampaignResponse:
    campaign = await get_campaign(session, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    await stop_campaign(campaign_id)
    await session.refresh(campaign)

    contacts = await get_campaign_contacts(session, campaign_id)
    return _to_response(campaign, contacts)


@router.get("/campaigns", response_model=list[CampaignResponse])
async def list_all(session: AsyncSession = Depends(get_session)) -> list[CampaignResponse]:
    campaigns = await list_campaigns(session)
    result = []
    for c in campaigns:
        contacts = await get_campaign_contacts(session, c.id)
        result.append(_to_response(c, contacts))
    return result


@router.get("/campaigns/{campaign_id}", response_model=CampaignResponse)
async def get_one(
    campaign_id: str, session: AsyncSession = Depends(get_session)
) -> CampaignResponse:
    campaign = await get_campaign(session, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    contacts = await get_campaign_contacts(session, campaign_id)
    return _to_response(campaign, contacts)


@router.get("/campaigns/{campaign_id}/export")
async def export_campaign_excel(
    campaign_id: str,
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Download campaign results as a formatted Excel file with call status and AI summaries."""
    campaign = await get_campaign(session, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    contacts = await get_campaign_contacts(session, campaign_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaign Results"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["#", "Name", "Phone Number", "Call Status", "Duration", "Called At", "AI Summary"]
    ws.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    # ── Status colours ────────────────────────────────────────────────────────
    STATUS_FILL = {
        "completed":   "D1FAE5",
        "in_progress": "DBEAFE",
        "voicemail":   "FEF3C7",
        "failed":      "FEE2E2",
        "no_answer":   "F3F4F6",
        "busy":        "FEE2E2",
        "cancelled":   "F3F4F6",
        "calling":     "EDE9FE",
        "pending":     "F9FAFB",
    }

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, contact in enumerate(contacts, 1):
        call = await get_call(session, contact.call_id) if contact.call_id else None

        raw_status = str(call.status if call else contact.status)
        # Human-friendly label
        label_map = {
            "completed": "Connected",
            "in_progress": "In Progress",
            "voicemail": "Voicemail",
            "failed": "Not Answered",
            "no_answer": "No Answer",
            "busy": "Busy",
            "cancelled": "Cancelled",
            "calling": "Calling",
            "pending": "Pending",
        }
        status_label = label_map.get(raw_status, raw_status.replace("_", " ").title())

        if call and call.duration_seconds:
            mins, secs = divmod(call.duration_seconds, 60)
            duration_str = f"{mins}:{secs:02d}"
        else:
            duration_str = "—"

        called_at = (
            call.created_at.strftime("%Y-%m-%d %H:%M") if call and call.created_at else "—"
        )
        summary = (call.summary or "") if call else ""

        ws.append([i, contact.name, contact.phone_number, status_label, duration_str, called_at, summary])

        row_num = i + 1
        color = STATUS_FILL.get(raw_status, "FFFFFF")
        ws.cell(row=row_num, column=4).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row_num, column=7).alignment = Alignment(wrap_text=True)

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, width in zip(range(1, 8), [5, 25, 18, 16, 10, 20, 70]):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = campaign.name.replace(" ", "_").replace("/", "_")[:40]
    filename = f"{safe_name}_results.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _to_response(campaign, contacts) -> CampaignResponse:
    return CampaignResponse(
        id=campaign.id,
        name=campaign.name,
        status=campaign.status,
        total_contacts=campaign.total_contacts,
        done_contacts=campaign.done_contacts or 0,
        created_at=campaign.created_at,
        contacts=[
            ContactStatusResponse(
                id=c.id,
                order_index=c.order_index,
                name=c.name,
                phone_number=c.phone_number,
                status=c.status,
                call_id=c.call_id,
            )
            for c in contacts
        ],
    )
