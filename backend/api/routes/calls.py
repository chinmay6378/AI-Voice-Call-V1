"""
REST endpoints for call management.

POST  /call/start             — Initiate an outbound AI call
POST  /call/end/{call_id}     — Hang up an active call
GET   /call/status/{call_id}  — Get call status
GET   /call/transcript/{call_id} — Get full call transcript
GET   /call/logs/{call_id}    — Get call event log
GET   /calls/active           — List current active call (if any)
GET   /health                 — Health check
"""
from __future__ import annotations

import io
import uuid
from typing import Annotated

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status  # noqa: F401
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from config.settings import Settings, get_settings
from services.signalwire.client import SignalWireClient
from database import (
    create_call,
    finalize_call,
    get_active_call,
    get_all_calls,
    get_call,
    get_calls_by_campaign,
    get_session,
    update_call_sid,
)
from database.models.call import CallStatus
from database.schemas.call import (
    CallLogResponse,
    CallStartedResponse,
    CallStatusResponse,
    CallTranscriptResponse,
    EndCallRequest,
    HealthResponse,
    StartCallRequest,
)
from services.livekit.room_manager import LiveKitRoomManager
from utils.logger import get_logger

logger = get_logger(__name__)
router = APIRouter(prefix="/call", tags=["calls"])


# ── Dependencies ──────────────────────────────────────────────────────────────

def get_livekit(settings: Annotated[Settings, Depends(get_settings)]) -> LiveKitRoomManager:
    return LiveKitRoomManager(settings)


def _check_credentials(settings: Settings) -> None:
    """Raise HTTP 400 with a clear list of missing credentials before we attempt a call."""
    missing: list[str] = []

    # Always required (STT / LLM / TTS)
    if not settings.groq_api_key:        missing.append("Groq API Key")
    if not settings.deepgram_api_key:    missing.append("Deepgram API Key")
    if not settings.elevenlabs_api_key:  missing.append("ElevenLabs API Key")

    # LiveKit is always needed (room + agent dispatch)
    if not settings.livekit_url:         missing.append("LiveKit URL")
    if not settings.livekit_api_key:     missing.append("LiveKit API Key")
    if not settings.livekit_api_secret:  missing.append("LiveKit API Secret")

    provider = (settings.telephony_provider or "livekit_sip").lower()
    if provider == "livekit_sip":
        if not settings.livekit_sip_trunk_id:
            missing.append("LiveKit SIP Trunk ID (for Vobiz)")
    elif provider == "signalwire":
        if not settings.signalwire_project_id: missing.append("SignalWire Project ID")
        if not settings.signalwire_api_token:  missing.append("SignalWire API Token")
        if not settings.signalwire_space_url:  missing.append("SignalWire Space URL")
        if not settings.signalwire_from_number: missing.append("SignalWire From Number")
        if not settings.app_base_url or "localhost" in settings.app_base_url:
            missing.append("App Base URL (must be public — SignalWire webhooks need to reach it)")

    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Cannot start call — missing required credentials.",
                "missing": missing,
                "hint": "Go to Settings and save the missing values.",
            },
        )


# ── POST /call/start ──────────────────────────────────────────────────────────

@router.post("/start", response_model=CallStartedResponse, status_code=status.HTTP_201_CREATED)
async def start_call(
    body: StartCallRequest,
    request: Request,
    session: Annotated[AsyncSession, Depends(get_session)],
    settings: Annotated[Settings, Depends(get_settings)],
    lk: Annotated[LiveKitRoomManager, Depends(get_livekit)],
) -> CallStartedResponse:
    """
    Initiate an outbound AI phone call via LiveKit SIP (Vobiz trunk).

    Flow:
      1. Guard: reject if another call is already active.
      2. Create LiveKit room.
      3. Persist call record in DB.
      4. Dispatch agent worker to LiveKit room.
      5. LiveKit dials the customer via Vobiz SIP trunk.
      6. Return call_id immediately — the rest is async.
    """
    # 0. Pre-flight: verify all required credentials are configured
    _check_credentials(settings)

    # 1. One-call-at-a-time guard
    active = await get_active_call(session)
    if active:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"A call is already active: {active.id} ({active.status}). End it before starting a new one.",
        )

    call_id = str(uuid.uuid4())
    room_name = f"call-{call_id[:8]}"

    # 2. Create LiveKit room
    try:
        await lk.create_room(room_name, call_id=call_id)
    except Exception as exc:
        logger.error("livekit.create_room_failed", call_id=call_id, error=str(exc))
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to create LiveKit room: {exc}",
        )

    # 3. Persist call record
    call = await create_call(
        session,
        customer_name=body.customer_name,
        phone_number=body.phone_number,
        livekit_room_name=room_name,
    )
    # Overwrite the auto-generated ID with our pre-computed one so URLs are consistent
    # (We keep call.id as-is from create_call; room_name references the first 8 chars)

    # 4. Dispatch LiveKit agent
    try:
        dispatch_id = await lk.dispatch_agent(
            room_name,
            call_id=call.id,
            customer_name=body.customer_name,
            phone_number=body.phone_number,
        )
        call.livekit_dispatch_id = dispatch_id
        await session.commit()
    except Exception as exc:
        logger.error("livekit.dispatch_failed", call_id=call.id, error=str(exc))
        await finalize_call(session, call.id, status=CallStatus.FAILED, error_message=str(exc))
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to dispatch agent: {exc}",
        )

    # 4.5. Pre-create SIP dispatch rule BEFORE dialing so it's active when AMD
    # completes (~6s later) and SignalWire sends the SIP INVITE to LiveKit.
    # Creating it here (not in the SWML webhook) gives LiveKit's routing layer
    # time to propagate the rule before the INVITE arrives.
    if settings.livekit_sip_uri and (settings.telephony_provider or "livekit_sip").lower() == "signalwire":
        try:
            rule_id = await lk.create_call_dispatch_rule(room_name)
            call.livekit_sip_rule_id = rule_id
            await session.commit()
            logger.info("call.dispatch_rule_created", call_id=call.id, rule_id=rule_id, room=room_name)
        except Exception as exc:
            logger.warning("call.dispatch_rule_create_failed", call_id=call.id, error=str(exc))

    # 5. Dial customer — branch on configured telephony provider
    provider = (settings.telephony_provider or "livekit_sip").lower()

    if provider == "signalwire":
        # SignalWire dials the number; on answer it hits /webhooks/swml/{call_id}
        # which returns SWML that bridges the call into the LiveKit room.
        sw = SignalWireClient(settings)
        base = settings.app_base_url.rstrip("/")
        try:
            call_sid = await sw.create_outbound_call(
                to=body.phone_number,
                swml_webhook_url=f"{base}/webhooks/swml/{call.id}",
                status_callback_url=f"{base}/webhooks/status/{call.id}",
                amd_callback_url=f"{base}/webhooks/amd/{call.id}",
            )
            await update_call_sid(session, call.id, call_sid)
            logger.info("call.initiated_signalwire", call_id=call.id, sid=call_sid, to=body.phone_number)
        except Exception as exc:
            logger.error("signalwire.call_failed", call_id=call.id, error=str(exc))
            await finalize_call(session, call.id, status=CallStatus.FAILED, error_message=str(exc))
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"Failed to initiate call via SignalWire: {exc}",
            )
    else:
        # LiveKit SIP — dials directly via Vobiz SIP trunk
        try:
            participant_id = await lk.create_sip_participant(
                room_name,
                phone_number=body.phone_number,
                customer_name=body.customer_name,
                call_id=call.id,
            )
            await update_call_sid(session, call.id, participant_id)
            logger.info("call.initiated_livekit_sip", call_id=call.id, participant_id=participant_id, to=body.phone_number)
        except Exception as exc:
            logger.error("livekit.sip_call_failed", call_id=call.id, error=str(exc))
            await finalize_call(session, call.id, status=CallStatus.FAILED, error_message=str(exc))
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail=f"Failed to initiate call via LiveKit SIP: {exc}",
            )

    return CallStartedResponse(
        call_id=call.id,
        status=CallStatus.DIALING,
        message=f"Outbound call to {body.phone_number} initiated via {provider}. call_id={call.id}",
    )


# ── POST /call/end/{call_id} ──────────────────────────────────────────────────

@router.post("/end/{call_id}", response_model=CallStatusResponse)
async def end_call(
    call_id: str,
    body: EndCallRequest | None = None,
    session: AsyncSession = Depends(get_session),
    lk: LiveKitRoomManager = Depends(get_livekit),
) -> CallStatusResponse:
    """Terminate an active call by call_id."""
    call = await get_call(session, call_id)
    if not call:
        raise HTTPException(status_code=404, detail=f"Call {call_id} not found")

    terminal = {CallStatus.COMPLETED, CallStatus.FAILED, CallStatus.CANCELLED}
    if call.status in terminal:
        raise HTTPException(
            status_code=409,
            detail=f"Call {call_id} is already in a terminal state: {call.status}",
        )

    # Delete LiveKit room — disconnects agent and SIP participant (customer)
    if call.livekit_room_name:
        await lk.delete_room(call.livekit_room_name)

    updated = await finalize_call(session, call_id, status=CallStatus.CANCELLED)
    return CallStatusResponse.model_validate(updated)


# ── GET /call/status/{call_id} ────────────────────────────────────────────────

@router.get("/status/{call_id}", response_model=CallStatusResponse)
async def get_call_status(
    call_id: str,
    session: AsyncSession = Depends(get_session),
) -> CallStatusResponse:
    call = await get_call(session, call_id)
    if not call:
        raise HTTPException(status_code=404, detail=f"Call {call_id} not found")
    return CallStatusResponse.model_validate(call)


# ── GET /call/transcript/{call_id} ───────────────────────────────────────────

@router.get("/transcript/{call_id}", response_model=CallTranscriptResponse)
async def get_transcript(
    call_id: str,
    session: AsyncSession = Depends(get_session),
) -> CallTranscriptResponse:
    call = await get_call(session, call_id)
    if not call:
        raise HTTPException(status_code=404, detail=f"Call {call_id} not found")
    return CallTranscriptResponse(
        call_id=call.id,
        customer_name=call.customer_name,
        status=call.status,
        transcript=call.get_transcript(),
        summary=call.summary,
    )


# ── GET /call/logs/{call_id} ──────────────────────────────────────────────────

@router.get("/logs/{call_id}", response_model=CallLogResponse)
async def get_logs(
    call_id: str,
    session: AsyncSession = Depends(get_session),
) -> CallLogResponse:
    call = await get_call(session, call_id)
    if not call:
        raise HTTPException(status_code=404, detail=f"Call {call_id} not found")
    return CallLogResponse(call_id=call.id, logs=call.get_logs())


# ── GET /calls/active ─────────────────────────────────────────────────────────

_active_router = APIRouter(prefix="/calls", tags=["calls"])


@_active_router.get("", response_model=list[CallStatusResponse])
async def list_all_calls(
    session: AsyncSession = Depends(get_session),
    campaign_id: str | None = Query(None),
) -> list[CallStatusResponse]:
    if campaign_id:
        calls = await get_calls_by_campaign(session, campaign_id)
    else:
        calls = await get_all_calls(session)
    return [CallStatusResponse.model_validate(c) for c in calls]


@_active_router.get("/active", response_model=CallStatusResponse | None)
async def get_active(session: AsyncSession = Depends(get_session)) -> CallStatusResponse | None:
    call = await get_active_call(session)
    if not call:
        return None
    return CallStatusResponse.model_validate(call)


# ── POST /call/inbound/setup ──────────────────────────────────────────────────

@router.post("/inbound/setup", status_code=status.HTTP_200_OK)
async def setup_inbound(
    settings: Annotated[Settings, Depends(get_settings)],
    lk: Annotated[LiveKitRoomManager, Depends(get_livekit)],
) -> dict:
    """
    One-time setup: (re)create the permanent inbound SIP dispatch rule in LiveKit.

    This rule uses Individual routing (one room per caller, prefix 'inbound-')
    and auto-dispatches the voice-call-agent so no explicit dispatch is needed
    when an inbound call arrives.

    Run this once after deploying, or again if you need to reset the rule.
    The rule persists across restarts — no need to call this on every boot.

    Prerequisites (do these in the LiveKit dashboard before calling this):
      1. SIP → Inbound Trunks → edit trunk → clear the Numbers field (accept all)
      2. Delete any manually-created dispatch rules so there are no conflicts
    """
    try:
        rule_id = await lk.create_inbound_dispatch_rule()
        return {
            "rule_id": rule_id,
            "message": "Inbound dispatch rule created. Configure SignalWire phone number webhooks as follows:",
            "signalwire_config": {
                "swml_webhook": f"{settings.app_base_url.rstrip('/')}/webhooks/inbound/swml",
                "status_callback": f"{settings.app_base_url.rstrip('/')}/webhooks/inbound/status",
            },
        }
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Failed to create inbound dispatch rule: {exc}",
        )


@_active_router.get("/export")
async def export_all_calls_excel(session: AsyncSession = Depends(get_session)) -> StreamingResponse:
    """Download all call records as a formatted Excel file with AI summaries."""
    calls = await get_all_calls(session, limit=5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Call Results"

    headers = ["#", "Name", "Phone Number", "Call Status", "Duration", "Date & Time", "AI Summary"]
    ws.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="1E3A8A")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    STATUS_FILL = {
        "completed":   "D1FAE5",
        "in_progress": "DBEAFE",
        "voicemail":   "FEF3C7",
        "failed":      "FEE2E2",
        "no_answer":   "F3F4F6",
        "busy":        "FEE2E2",
        "cancelled":   "F3F4F6",
        "dialing":     "EDE9FE",
        "ringing":     "EDE9FE",
        "pending":     "F9FAFB",
    }
    label_map = {
        "completed": "Connected", "in_progress": "In Progress",
        "voicemail": "Voicemail", "failed": "Not Answered",
        "no_answer": "No Answer", "busy": "Busy",
        "cancelled": "Cancelled", "dialing": "Dialing",
        "ringing": "Ringing", "pending": "Pending",
    }

    for i, call in enumerate(calls, 1):
        raw_status = str(call.status)
        status_label = label_map.get(raw_status, raw_status.replace("_", " ").title())

        if call.duration_seconds:
            mins, secs = divmod(call.duration_seconds, 60)
            duration_str = f"{mins}:{secs:02d}"
        else:
            duration_str = "—"

        date_str = call.created_at.strftime("%Y-%m-%d %H:%M") if call.created_at else "—"
        summary = call.summary or ""

        ws.append([i, call.customer_name, call.phone_number, status_label, duration_str, date_str, summary])

        row_num = i + 1
        color = STATUS_FILL.get(raw_status, "FFFFFF")
        ws.cell(row=row_num, column=4).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row_num, column=7).alignment = Alignment(wrap_text=True)

    for col, width in zip(range(1, 8), [5, 25, 18, 16, 10, 20, 70]):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="call_results.xlsx"'},
    )
